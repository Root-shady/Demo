from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout

# Create your views here.
from rest_framework.decorators import api_view, renderer_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.renderers import JSONRenderer, BrowsableAPIRenderer
from rest_framework.parsers import JSONParser, FileUploadParser
from rest_framework.views import APIView
from .models import UserProfile
from .serializer import UserProfileSerializer
import openpyxl
from openpyxl.cell import  get_column_letter, column_index_from_string

@api_view(['GET', 'POST'])
def register(request, format=None):
	"""
	Create a new user
	"""
	# The GET method is only for debug purpose
	if request.method == 'GET':
		users = UserProfile.objects.all()
		serializer = UserProfileSerializer(users, many=True)
		return Response(serializer.data)
	elif request.method == 'POST':
		serializer = UserProfileSerializer(data=request.data)
		if serializer.is_valid():
			serializer.save()
			return Response(serializer.data, status=status.HTTP_201_CREATED)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@renderer_classes((BrowsableAPIRenderer,JSONRenderer,))
def user_login(request, format=None):
	"""
	Login a user
	"""
	if request.method == 'POST':
		data = request.data
		password = data['password']
		username = data['username']
		user = authenticate(username=username, password=password)
		feedback = {}
		if user is not None:
			# the password verified for the user
			if user.is_active:
				login(request, user)
				print(request.user)
				user = User.objects.get(username=username)
				user = UserProfile.objects.get(user=user)

				feedback['id'] = user.id
				feedback['is_admin'] = user.is_admin
				feedback['username'] = user.user.username
				# Handle the <ImageFieldFile: None> problem, ugly hack
				if user.image == '':
					feedback['image'] = None
				else:
					feedback['image'] = user.image
				return Response(feedback, status=status.HTTP_200_OK)
			else:
				feedback['error'] = "The user's account is being disabled"
				return Response(feedback, status=status.HTTP_400_BAD_REQUEST)
		else:
			feedback['error'] = "Username and password don't match"
			return Response(feedback, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def user_logout(request, format=None):
	logout(request)
	return Response(status=status.HTTP_200_OK)

@api_view(['GET'])
def user_profile(request, user_id, format=None):
	"""
	Display specific user information
	"""
	print(request.user)
	if not request.user.is_authenticated():
		return Response(status=status.HTTP_400_BAD_REQUEST)
	if request.method == 'GET':
		try:
			user = UserProfile.objects.get(pk=user_id)
		except UserProfile.DoesNotExist:
			return Response(status=status.HTTP_404_NOT_FOUND)
		serializer = UserProfileSerializer(user)
		return Response(serializer.data)

@api_view(['POST'])
def user_edit(request, user_id, format=None):
	"""
	Update the user profile
	"""
	print(request.user.username)
	print(request.user.id)
	if not request.user.is_authenticated or request.user.id != user_id:
		return Response(status=status.HTTP_400_BAD_REQUEST)

	try:
		user = UserProfile.objects.get(pk=user_id)
	except UserProfile.DoesNotExist:
		return Response(status=status.HTTP_404_NOT_FOUND)
	if request.method == 'POST':
		serializer =  UserProfileSerializer(user, data=request.data)
		if serializer.is_valid():
			serializer.save()
			return Response(serializer.data)
		return Response(serializer.errors, status.HTTP_400_BAD_REQUEST)

# The below is the admin method
#def is_authenticated(request):

@api_view(['GET'])
def admin_all(request, format=None):
	"""
	Return all the created user to the admin
	"""
	users = UserProfile.objects.all()
	if request.method == 'GET':
		serializer = UserProfileSerializer(users, many=True)
		return Response(serializer.data)

@api_view(['DELETE', 'GET'])
def admin_del(request, user_id, format=None):
	"""
	Delete specified user
	"""
	try:
		userPro = UserProfile.objects.get(pk=user_id)
	except UserProfile.DoesNotExist:
		return Response(status=status.HTTP_404_NOT_FOUND)
	if request.method == 'GET':
		serializer = UserProfileSerializer(userPro)
		return Response(serializer.data)
	if request.method == 'DELETE':
		userPro.user.delete()
		userPro.delete()
		return Response(status=status.HTTP_200_OK)

@api_view(['GET', 'POST'])
@renderer_classes((BrowsableAPIRenderer,JSONRenderer,))
def admin_update(request, user_id, format=None):
	"""
	Update the user profile
	"""
	try:
		userPro = UserProfile.objects.get(pk=user_id)
	except UserProfile.DoesNotExist:
		return Response(status=status.HTTP_404_NOT_FOUND)
	if request.method == 'GET':
		serializer = UserProfileSerializer(userPro)
		return Response(serializer.data)
	# User input data validation
	if request.method == 'POST':
		#serializer =  UserProfileSerializer(user, data=request.data)
		try:
			data = request.data
			userPro.sex = data.get('sex', userPro.sex)
			userPro.image = data.get('image', userPro.image)

			if 'user' in data:
				# Username validation
				username = data['user'].get('username', userPro.user.username)
				if userPro.user.username != username:
					try:
						valid_check = User.objects.get(username=username)
						data = {'error': 'An username has existed'}
						return Response(data, status=status.HTTP_400_BAD_REQUEST)
					except:
						userPro.user.username = username
				userPro.user.set_password(data['user'].get('password', userPro.user.password))
				# Email Validation
				userPro.user.email = data['user'].get('email', userPro.user.email)
			# Save the user in the user model
			userPro.user.save()
			userPro.save()
			return  Response(status.status.HTTP_200_OK)
		except:
			return Response(serializer.errors, status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def admin_userAdd(request, format=None):
	"""
	Take in an excel file, than read the file and create user
	"""
	wb =  openpyxl.load_workbook('book1.xlsx')
	sheet =  wb.get_sheet_by_name('Sheet1')


# Admin can generated excel file
@api_view(['GET'])
def admin_excel(request, format=None):
	"""
	When the admin request for the excel file, it generate a excel file 
	according to the current system
	"""
	if request.method == 'GET':
		wb = openpyxl.Workbook()
		sheet = wb.active
		sheet.title = "Sheet1"
		sheet['A1'] = 'USERNAME'
		sheet['B1'] = 'PASSWORD'
		sheet['C1'] = 'EMAIL'
		sheet['D1'] = 'SEX'
		users = UserProfile.objects.all()
		count = 2
		for user in users:
			sheet['A'+str(count)] = user.username
			sheet['B'+str(count)] = user.password
			sheet['C'+str(count)] = user.email
			sheet['D'+str(count)] = user.sex
			count += 1
		wb.save('../collectStatic/media_root/user.xlsx')
		data = {"download_url":"media/user.xlsx"}
		return Response(data, status.HTTP_200_OK)

@api_view(['POST'])
def upload_form(request):
	if request.method == 'POST':
		instance = Files(docfile=request.FILES['docfile'], title=request.DATA['title'])
		instance.save()
		return Response('uploaded')


@api_view(['POST'])
@renderer_classes((JSONRenderer, ))
def admin_upload(request, format='None'):
	parser_classes = (FileUploadParser, )
	up_file = request.FILES.get('file', None)
	if up_file is None:
		return Response("Emppty")
	print(up_file.name)
	path = '../collectStatic/media_root/' + up_file
	destination = open(path, 'wb+')
	for chunk in up_file.chunks():
		destination.write(chunk)
		destination.close()
	# ...
	# do some stuff with uploaded file
	# ...
	with open(path, 'rb') as fp:
		content = fp.readlines()
		print(content)
 # do some stuff with uploaded file
		wb = openpyxl.load_workbook(path)
		sheet = wb.get_sheet_by_name('Sheet1')
		print(sheet.max_column)
		print(sheet.max_row)
		endpoint = get_column_letter(sheet.max_column) + str(sheet.max_row)
		if sheet.max_column < 3:
				data['error'] = 'There should be username, email, password'
				return Response(data, status=status.HTTP_400_BAD_REQUEST)
		for rows in sheet['A2':endpoint]:
				username, email, password = rows[0].value, rows[1].value, rows[2].value
				if password is None or password == '':
						password = 'password'
				# Try to create user, if existed, then add the user name to the list
				try:
						user = UserProfile(username=username, password=password, email=email)
						user.save()
				except:
						data['user_list'].append(username)
						print(data['user_list'])
		#response['status'] = status.HTTP_200_OK
		return Response(data, status.HTTP_200_OK)

